import pytest

from app.services.tracking_plan import (
    attach_property,
    create_destination,
    create_event,
    create_property,
    create_source,
    get_main_branch,
    get_or_create_plan,
    plan_to_dict,
    set_event_destination,
    set_event_sources,
)
from app.services.tracking_plan.exports import plan_to_markdown, plan_to_xlsx
from tests.services.tracking_plan.test_models import _make_project_and_user


async def _sample_plan_dict(session):
    project_id, user_id = await _make_project_and_user(session)
    plan = await get_or_create_plan(session, project_id=project_id, user_id=user_id, name="P")
    branch = await get_main_branch(session, plan)
    ev = await create_event(session, branch, name="purchase", purpose="completes checkout")
    prop = await create_property(session, branch, name="value", data_type="float")
    await attach_property(session, branch, ev.id, prop.id, required=True, example="9.99")
    src = await create_source(session, branch, name="web", platform_type="web")
    dest = await create_destination(session, branch, name="GA4", platform="ga4")
    await set_event_sources(
        session, branch, ev.id, [{"source_id": src.id, "implementation_status": "implemented"}]
    )
    await set_event_destination(session, branch, ev.id, dest.id, dest_event_name="purchase")
    return await plan_to_dict(session, plan, branch)


@pytest.mark.anyio
async def test_plan_to_markdown(db_session_factory):
    async with db_session_factory() as session:
        data = await _sample_plan_dict(session)
        md = plan_to_markdown(data)
        assert "# P" in md
        assert "## purchase" in md
        assert "value" in md
        assert "9.99" in md
        assert "GA4" in md


@pytest.mark.anyio
async def test_plan_to_xlsx_is_valid_workbook(db_session_factory):
    import io

    from openpyxl import load_workbook

    async with db_session_factory() as session:
        data = await _sample_plan_dict(session)
        raw = plan_to_xlsx(data)
        wb = load_workbook(io.BytesIO(raw))
        assert "Events" in wb.sheetnames
        assert "Properties" in wb.sheetnames
        # Events sheet has a header row + the purchase row
        ws = wb["Events"]
        names = [c.value for c in ws["A"]]
        assert "purchase" in names
