# tests/test_sdr_richer_sections.py
"""Parser + export tests for the richer SDR sections."""

from app.tools.sdr_parser import parse_markdown_table, parse_sdr_markdown

SAMPLE = """---
sdr_name: Demo SDR
---

# Demo SDR

## Executive Summary

| Property | GA4 123 (demo.com) |
|---|---|
| GTM Container | GTM-XXX |
| Platforms in scope | GA4, GTM, Google Ads |

Short narrative summary here.

---

## Gap Register

| # | Severity | Finding | Evidence | Business impact | Recommended fix | Fix location | Owner |
|---|---|---|---|---|---|---|---|
| 1 | Critical | No primary conversion | split events | no ROAS | one generate_lead | data layer | Analytics |

---

## Event Catalog

### `demo_confirmation`

*Status:* `implemented` | *Last verified:* `never`

**Business Purpose:** Demo booked.

**Triggers:**
- Type: `datalayer_event`
- Configuration: on confirm

**Destinations:**

- **GA4**: event name `demo_confirmation`

---

## Conversion Audit

| GA4 key event | 90d count | Unique converters | Fires? | Verdict / action |
|---|---|---|---|---|
| content_click | 135201 | 25719 | Yes | REMOVE as conversion |

---

## Consent & Privacy

| # | Severity | Check | Finding | Recommendation |
|---|---|---|---|---|
| 1 | Critical | Consent Mode v2 | missing | add defaults |

---

## Remediation Roadmap

| Phase | Action | Resolves | Effort | Impact | Owner |
|---|---|---|---|---|---|
| Phase 1 | Un-mark content_click | #2 | Low | Very High | Analytics |

---

## Changelog
"""


def test_parser_populates_new_raw_sections():
    parsed = parse_sdr_markdown(SAMPLE)
    assert parsed.executive_summary and "GTM Container" in parsed.executive_summary
    assert parsed.gap_register and "No primary conversion" in parsed.gap_register
    assert parsed.conversion_audit and "content_click" in parsed.conversion_audit
    assert parsed.consent_and_privacy and "Consent Mode v2" in parsed.consent_and_privacy
    assert parsed.remediation_roadmap and "Un-mark content_click" in parsed.remediation_roadmap


def test_new_sections_parse_into_tables():
    parsed = parse_sdr_markdown(SAMPLE)
    gap = parse_markdown_table(parsed.gap_register)
    assert gap["headers"][1] == "Severity"
    assert gap["rows"][0][2] == "No primary conversion"
    roadmap = parse_markdown_table(parsed.remediation_roadmap)
    assert roadmap["headers"] == ["Phase", "Action", "Resolves", "Effort", "Impact", "Owner"]


def test_unrelated_headings_not_misrouted():
    md = (
        "# X\n\n"
        "## Key Findings\n\nsome prose\n\n---\n\n"
        "## Technical Roadmap\n\nmore prose\n\n---\n\n"
        "## Event Catalog\n\n### `e`\n\n*Status:* `planned`\n\n"
        "**Business Purpose:** x\n\n**Triggers:**\n- Type: `click`\n"
    )
    parsed = parse_sdr_markdown(md)
    # 'Key Findings' must NOT be captured as the structured Gap Register
    assert parsed.gap_register is None
    # 'Technical Roadmap' must NOT be captured as Remediation Roadmap
    assert parsed.remediation_roadmap is None


def test_export_includes_new_sheets():
    from io import BytesIO

    from openpyxl import load_workbook

    from app.tools.sdr_excel_export import generate_sdr_xlsx

    xlsx = generate_sdr_xlsx(SAMPLE, "Demo SDR")
    wb = load_workbook(BytesIO(xlsx))
    for name in ["Executive Summary", "Gap Register", "Conversion Audit",
                 "Consent & Privacy", "Remediation Roadmap"]:
        assert name in wb.sheetnames, f"missing sheet {name}"
    ws = wb["Gap Register"]
    assert ws.cell(row=1, column=1).value == "#"
    assert ws.cell(row=2, column=3).value == "No primary conversion"


def test_export_omits_absent_sections():
    from io import BytesIO

    from openpyxl import load_workbook

    from app.tools.sdr_excel_export import generate_sdr_xlsx

    minimal = "# X\n\n## Event Catalog\n\n### `e`\n\n*Status:* `planned`\n\n**Business Purpose:** x\n\n**Triggers:**\n- Type: `click`\n"
    wb = load_workbook(BytesIO(generate_sdr_xlsx(minimal, "X")))
    assert "Gap Register" not in wb.sheetnames
