# app/services/tracking_plan/exports.py
"""Generate human artifacts (markdown, xlsx) from a plan_to_dict() dict.

These are OUTPUTS only — the relational tables remain the source of truth."""

import io

from openpyxl import Workbook


def plan_to_markdown(plan: dict) -> str:
    """Render a readable Markdown tracking-plan doc from a plan_to_dict() dict."""
    lines: list[str] = []
    lines.append(f"# {plan['plan']['name']}")
    if plan["plan"].get("description"):
        lines.append("")
        lines.append(plan["plan"]["description"])
    lines.append("")

    lines.append("## Events")
    lines.append("")
    for ev in plan["events"]:
        lines.append(f"## {ev['name']}")
        if ev.get("display_name"):
            lines.append(f"*{ev['display_name']}*")
        if ev.get("category"):
            lines.append(f"- **Category:** {ev['category']}")
        if ev.get("purpose"):
            lines.append(f"- **Purpose:** {ev['purpose']}")
        if ev.get("trigger_type"):
            lines.append(f"- **Trigger:** {ev['trigger_type']}")
        if ev.get("tags"):
            lines.append(f"- **Tags:** {', '.join(ev['tags'])}")
        if ev["sources"]:
            srcs = ", ".join(f"{s['name']} ({s['implementation_status']})" for s in ev["sources"])
            lines.append(f"- **Sources:** {srcs}")
        lines.append("")
        if ev["properties"]:
            lines.append("| Property | Type | Required | Example |")
            lines.append("| --- | --- | --- | --- |")
            for p in ev["properties"]:
                lines.append(
                    f"| {p['name']} | {p['data_type']} | {'yes' if p['required'] else 'no'} | {p.get('example') or ''} |"
                )
            lines.append("")
        if ev["destinations"]:
            lines.append("**Destinations:**")
            for d in ev["destinations"]:
                lines.append(f"- {d['destination']}: `{d.get('dest_event_name') or ev['name']}`")
            lines.append("")

    if plan["properties"]["user"]:
        lines.append("## User Properties")
        lines.append("")
        lines.append("| Name | Type |")
        lines.append("| --- | --- |")
        for p in plan["properties"]["user"]:
            lines.append(f"| {p['name']} | {p['data_type']} |")
        lines.append("")

    if plan["sources"]:
        lines.append("## Sources → Destinations")
        lines.append("")
        for s in plan["sources"]:
            routed = ", ".join(s["destinations"]) or "—"
            lines.append(f"- **{s['name']}** ({s.get('platform_type') or 'n/a'}) → {routed}")
        lines.append("")

    if plan["metrics"]:
        lines.append("## Metrics")
        lines.append("")
        for m in plan["metrics"]:
            lines.append(f"- **{m['name']}** — {m.get('event') or 'n/a'}")
            if m.get("description"):
                lines.append(f"  {m['description']}")
        lines.append("")

    return "\n".join(lines).rstrip() + "\n"


def plan_to_xlsx(plan: dict) -> bytes:
    """Render a multi-sheet workbook from a plan_to_dict() dict."""
    wb = Workbook()

    ws = wb.active
    ws.title = "Events"
    ws.append(["name", "display_name", "category", "purpose", "trigger_type", "sources", "destinations"])
    for ev in plan["events"]:
        ws.append(
            [
                ev["name"],
                ev.get("display_name") or "",
                ev.get("category") or "",
                ev.get("purpose") or "",
                ev.get("trigger_type") or "",
                "; ".join(f"{s['name']}:{s['implementation_status']}" for s in ev["sources"]),
                "; ".join(d["destination"] for d in ev["destinations"]),
            ]
        )

    wp = wb.create_sheet("Properties")
    wp.append(["event", "property", "data_type", "required", "example"])
    for ev in plan["events"]:
        for p in ev["properties"]:
            wp.append(
                [
                    ev["name"],
                    p["name"],
                    p["data_type"],
                    "yes" if p["required"] else "no",
                    p.get("example") or "",
                ]
            )
    for p in plan["properties"]["user"]:
        wp.append(["(user)", p["name"], p["data_type"], "", ""])

    wd = wb.create_sheet("Destinations")
    wd.append(["name", "platform", "account_id"])
    for d in plan["destinations"]:
        wd.append([d["name"], d["platform"], d.get("platform_account_id") or ""])

    ws_src = wb.create_sheet("Sources")
    ws_src.append(["name", "platform_type", "routes_to"])
    for s in plan["sources"]:
        ws_src.append([s["name"], s.get("platform_type") or "", "; ".join(s["destinations"])])

    wm = wb.create_sheet("Metrics")
    wm.append(["name", "event", "description"])
    for m in plan["metrics"]:
        wm.append([m["name"], m.get("event") or "", m.get("description") or ""])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
