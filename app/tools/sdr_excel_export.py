"""
SDR Excel Export — generate a multi-sheet .xlsx from parsed SDR data.

Produces an industry-standard Solution Design Reference spreadsheet with:
  Sheet 1: Event Catalog      — one row per event
  Sheet 2: Event Parameters    — one row per parameter per event
  Sheet 3: Destinations Matrix — cross-reference (events × platforms)
  Sheet 4: User Properties     — custom dimensions
  Sheet 5: Business Context    — metadata key-value pairs

Styled with conditional formatting: status colors, TODO highlights,
required badges, frozen headers, auto-width columns.
"""

from __future__ import annotations

import io
import logging

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from app.tools.sdr_parser import ParsedSDR, parse_sdr_markdown

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

_HEADER_FONT = Font(name="Inter", size=10, bold=True, color="3C3836")
_HEADER_FILL = PatternFill(start_color="F5F3ED", end_color="F5F3ED", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=False)
_CELL_FONT = Font(name="Inter", size=10, color="1C1917")
_MONO_FONT = Font(name="JetBrains Mono", size=9, color="1C1917")
_MUTED_FONT = Font(name="Inter", size=10, color="78716C")
_THIN_BORDER = Border(
    bottom=Side(style="thin", color="E4E2DC"),
)

# Status fills
_STATUS_FILLS = {
    "planned": PatternFill(start_color="F5F5F4", end_color="F5F5F4", fill_type="solid"),
    "implemented": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
    "verified": PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
    "deprecated": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
}
_STATUS_FONTS = {
    "planned": Font(name="Inter", size=9, bold=True, color="78716C"),
    "implemented": Font(name="Inter", size=9, bold=True, color="2563EB"),
    "verified": Font(name="Inter", size=9, bold=True, color="16A34A"),
    "deprecated": Font(name="Inter", size=9, bold=True, color="DC2626"),
}

_TODO_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
_TODO_FONT = Font(name="Inter", size=9, bold=True, color="D97706")

_REQUIRED_YES_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
_REQUIRED_YES_FONT = Font(name="Inter", size=9, bold=True, color="16A34A")
_REQUIRED_NO_FONT = Font(name="Inter", size=9, color="A8A29E")

_CHECK_FONT = Font(name="Inter", size=11, bold=True, color="16A34A")
_DASH_FONT = Font(name="Inter", size=11, color="D6D3D1")

_LABEL_FONT = Font(name="Inter", size=9, bold=True, color="78716C")
_VALUE_FONT = Font(name="Inter", size=10, color="1C1917")


def _is_todo(val: str | None) -> bool:
    """Check if a value is a TODO marker."""
    return bool(val and "[TODO" in val)


def _clean(val: str | None, max_len: int = 500) -> str:
    """Clean a value for Excel: strip TODO markers, truncate."""
    if not val:
        return ""
    # Strip [TODO: ...] markers for cleaner display
    s = val.strip()
    if len(s) > max_len:
        s = s[:max_len] + "…"
    return s


def _auto_width(ws, min_width: int = 10, max_width: int = 50) -> None:
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, len(val))
            except Exception:
                pass
        adjusted = min(max(max_len + 3, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def _write_header(ws, headers: list[str], row: int = 1) -> None:
    """Write styled header row."""
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------


def _build_event_catalog(wb: Workbook, parsed: ParsedSDR) -> None:
    """Sheet 1: Event Catalog — one row per event."""
    ws = wb.active
    ws.title = "Event Catalog"

    headers = [
        "#",
        "Event Name",
        "Business Purpose",
        "Trigger Type",
        "Trigger Config",
        "Status",
        "Params Count",
        "Destinations",
        "Owner (Business)",
        "Owner (Technical)",
        "Consent",
        "Related KPIs",
    ]
    _write_header(ws, headers)
    ws.freeze_panes = "A2"

    for idx, event in enumerate(parsed.events, 1):
        row = idx + 1
        status = event.status or "planned"
        dest_str = ", ".join(d.platform.upper() for d in event.destinations) if event.destinations else ""
        consent_str = " | ".join(event.consent_required) if event.consent_required else ""
        kpi_str = ", ".join(event.kpi_links) if event.kpi_links else ""

        trigger_config_str = ""
        if event.trigger_config:
            parts = []
            for k, v in event.trigger_config.items():
                parts.append(f"{k}: {v}")
            trigger_config_str = "; ".join(parts)

        values = [
            idx,
            event.name,
            _clean(event.purpose),
            event.trigger_type or "",
            trigger_config_str,
            status.upper(),
            len(event.parameters),
            dest_str,
            event.owner_business or "",
            event.owner_technical or "",
            consent_str,
            kpi_str,
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = _CELL_FONT
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(col in (3, 5, 12)))

        # Style event name as mono
        ws.cell(row=row, column=2).font = _MONO_FONT

        # Style status cell
        status_cell = ws.cell(row=row, column=6)
        if status in _STATUS_FILLS:
            status_cell.fill = _STATUS_FILLS[status]
            status_cell.font = _STATUS_FONTS[status]
            status_cell.alignment = Alignment(horizontal="center", vertical="top")

        # Highlight TODO purposes
        purpose_cell = ws.cell(row=row, column=3)
        if _is_todo(event.purpose):
            purpose_cell.fill = _TODO_FILL
            purpose_cell.font = _TODO_FONT
            purpose_cell.value = (
                "[TODO] " + _clean(event.purpose).replace("[TODO:", "").replace("[TODO]", "").strip()
            )

        # Highlight missing triggers
        trigger_cell = ws.cell(row=row, column=4)
        if not event.trigger_type:
            trigger_cell.fill = _TODO_FILL
            trigger_cell.font = _TODO_FONT
            trigger_cell.value = "[TODO]"

    _auto_width(ws)


def _build_parameters(wb: Workbook, parsed: ParsedSDR) -> None:
    """Sheet 2: Event Parameters — one row per parameter per event."""
    ws = wb.create_sheet("Parameters")

    headers = [
        "Event Name",
        "Parameter",
        "Type",
        "Required",
        "Source",
        "Example",
        "Validation Rule",
    ]
    _write_header(ws, headers)
    ws.freeze_panes = "A2"

    row = 2
    for event in parsed.events:
        for param in event.parameters:
            values = [
                event.name,
                param.name,
                param.type or "",
                "YES" if param.required else "NO",
                param.source or "",
                param.example or "",
                param.validation_rule or "",
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = _CELL_FONT
                cell.border = _THIN_BORDER
                cell.alignment = Alignment(vertical="top")

            # Mono font for event + param names
            ws.cell(row=row, column=1).font = _MONO_FONT
            ws.cell(row=row, column=2).font = _MONO_FONT

            # Style required cell
            req_cell = ws.cell(row=row, column=4)
            req_cell.alignment = Alignment(horizontal="center", vertical="top")
            if param.required:
                req_cell.fill = _REQUIRED_YES_FILL
                req_cell.font = _REQUIRED_YES_FONT
            else:
                req_cell.font = _REQUIRED_NO_FONT

            # Highlight TODO type
            type_cell = ws.cell(row=row, column=3)
            if not param.type or _is_todo(param.type):
                type_cell.fill = _TODO_FILL
                type_cell.font = _TODO_FONT
                type_cell.value = "[TODO]"

            # Source mono
            ws.cell(row=row, column=5).font = _MONO_FONT
            # Example mono
            ws.cell(row=row, column=6).font = _MONO_FONT

            row += 1

    _auto_width(ws)


def _build_destinations_matrix(wb: Workbook, parsed: ParsedSDR) -> None:
    """Sheet 3: Destinations Matrix — events × platforms cross-reference."""
    ws = wb.create_sheet("Destinations")

    # Collect all platforms
    all_platforms: list[str] = []
    for event in parsed.events:
        for d in event.destinations:
            if d.platform not in all_platforms:
                all_platforms.append(d.platform)

    # Ensure common platforms appear even if unused
    for plat in ["ga4", "google_ads", "meta", "tiktok", "snap", "linkedin", "pinterest"]:
        if plat not in all_platforms:
            all_platforms.append(plat)

    headers = ["Event Name", "Status"] + [p.upper() for p in all_platforms]
    _write_header(ws, headers)
    ws.freeze_panes = "A2"

    for idx, event in enumerate(parsed.events, 1):
        row = idx + 1
        status = event.status or "planned"

        # Event name
        name_cell = ws.cell(row=row, column=1, value=event.name)
        name_cell.font = _MONO_FONT
        name_cell.border = _THIN_BORDER

        # Status
        status_cell = ws.cell(row=row, column=2, value=status.upper())
        status_cell.border = _THIN_BORDER
        if status in _STATUS_FILLS:
            status_cell.fill = _STATUS_FILLS[status]
            status_cell.font = _STATUS_FONTS[status]
            status_cell.alignment = Alignment(horizontal="center")

        # Platform columns
        event_dests = {d.platform: d for d in event.destinations}
        for col_offset, plat in enumerate(all_platforms):
            cell = ws.cell(row=row, column=3 + col_offset)
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

            if plat in event_dests:
                dest = event_dests[plat]
                mapped = dest.dest_event_name
                if mapped and mapped != event.name:
                    cell.value = f"✓ ({mapped})"
                else:
                    cell.value = "✓"
                cell.font = _CHECK_FONT
            else:
                cell.value = "—"
                cell.font = _DASH_FONT

    _auto_width(ws, min_width=12)


def _build_user_properties(wb: Workbook, parsed: ParsedSDR) -> None:
    """Sheet 4: User Properties / Custom Dimensions."""
    ws = wb.create_sheet("User Properties")

    headers = ["Property Name", "Scope", "Source", "Example Values", "Platforms"]
    _write_header(ws, headers)
    ws.freeze_panes = "A2"

    uprops = parsed.user_properties or ""
    if uprops and "[TODO" not in uprops:
        # Try to parse as a markdown table
        lines = [ln.strip() for ln in uprops.split("\n") if ln.strip().startswith("|")]
        data_lines = [ln for ln in lines if not ln.startswith("|---") and not ln.startswith("| ---")]
        if len(data_lines) >= 2:
            # Skip header row, parse data
            for line in data_lines[1:]:
                cells = [c.strip() for c in line.split("|") if c.strip()]
                row = ws.max_row + 1
                for col, val in enumerate(cells[:5], 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.font = _CELL_FONT
                    cell.border = _THIN_BORDER
        else:
            # Free-form text — put it in a note
            cell = ws.cell(row=2, column=1, value=uprops[:500])
            cell.font = _MUTED_FONT
            cell.alignment = Alignment(wrap_text=True)
    else:
        cell = ws.cell(
            row=2, column=1, value="No user properties documented yet. Refine the SDR via Claude to add them."
        )
        cell.font = _MUTED_FONT

    _auto_width(ws)


def _build_business_context(wb: Workbook, parsed: ParsedSDR) -> None:
    """Sheet 5: Business Context & Metadata — key-value layout."""
    ws = wb.create_sheet("Business Context")

    headers = ["Section", "Content"]
    _write_header(ws, headers)
    ws.freeze_panes = "A2"

    last_updated = parsed.last_updated
    if last_updated and not isinstance(last_updated, str):
        last_updated = str(last_updated)

    sections = [
        ("Business Type", parsed.business_type or "—"),
        ("SDR Status", parsed.sdr_status or "—"),
        ("SDR Version", parsed.sdr_version or "—"),
        ("Last Updated", last_updated or "—"),
        ("", ""),  # spacer
        ("Business Context", _clean(parsed.business_context, 2000) or "[TODO]"),
        ("User Journeys", _clean(parsed.user_journeys, 2000) or "[TODO]"),
        ("Data Layer Schema", _clean(parsed.data_layer_schema, 2000) or "[TODO]"),
        ("Consent & Privacy", _clean(parsed.consent_and_privacy, 2000) or "[TODO]"),
        ("Ownership & Governance", _clean(parsed.ownership, 2000) or "[TODO]"),
    ]

    for idx, (label, value) in enumerate(sections, 2):
        label_cell = ws.cell(row=idx, column=1, value=label)
        label_cell.font = _LABEL_FONT
        label_cell.alignment = Alignment(vertical="top")
        label_cell.border = _THIN_BORDER

        value_str = str(value) if value is not None else ""
        value_cell = ws.cell(row=idx, column=2, value=value_str)
        value_cell.font = _VALUE_FONT
        value_cell.alignment = Alignment(vertical="top", wrap_text=True)
        value_cell.border = _THIN_BORDER

        # Highlight TODOs
        if "[TODO" in value_str:
            value_cell.fill = _TODO_FILL
            value_cell.font = _TODO_FONT

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 80


def _build_table_sheet(wb: Workbook, sheet_title: str, raw_markdown: str | None) -> None:
    """Generic sheet: render a markdown table section as headers + rows.

    No-op (no sheet created) when the section is absent or has no table — keeps
    the workbook free of empty sheets.
    """
    from app.tools.sdr_parser import parse_markdown_table

    table = parse_markdown_table(raw_markdown)
    if not table or not table.get("rows"):
        return

    ws = wb.create_sheet(sheet_title)
    _write_header(ws, table["headers"])
    ws.freeze_panes = "A2"
    for line in table["rows"]:
        row = ws.max_row + 1
        for col, val in enumerate(line, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = _CELL_FONT
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    _auto_width(ws)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def generate_sdr_xlsx(markdown_content: str, sdr_name: str = "SDR") -> bytes:
    """
    Generate a multi-sheet .xlsx from SDR markdown content.

    Returns the Excel file as bytes (ready for StreamingResponse).
    """
    parsed = parse_sdr_markdown(markdown_content)

    wb = Workbook()

    _build_table_sheet(wb, "Executive Summary", parsed.executive_summary)
    _build_event_catalog(wb, parsed)
    _build_parameters(wb, parsed)
    _build_table_sheet(wb, "Conversion Audit", parsed.conversion_audit)
    _build_user_properties(wb, parsed)
    _build_destinations_matrix(wb, parsed)
    _build_table_sheet(wb, "Consent & Privacy", parsed.consent_and_privacy)
    _build_table_sheet(wb, "Gap Register", parsed.gap_register)
    _build_table_sheet(wb, "Remediation Roadmap", parsed.remediation_roadmap)
    _build_business_context(wb, parsed)

    # Write to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
